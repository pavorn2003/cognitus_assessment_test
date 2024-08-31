from io import BytesIO
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse
import uvicorn
import pandas as pd
import numpy as np
import re

from openpyxl import load_workbook
from pandas.api.types import is_numeric_dtype
# from gensim.models import KeyedVectors
from scipy.spatial.distance import cosine
from sklearn.ensemble import IsolationForest
from scipy.stats import zscore
import fasttext
import fasttext.util
from pydantic import BaseModel
from typing import List, Tuple
import fasttext
import requests
from tqdm import tqdm
import os

model_url = "https://dl.fbaipublicfiles.com/fasttext/vectors-crawl/cc.en.300.bin.gz"
model_path = "cc.en.300.bin.gz"
extracted_model_path = "cc.en.300.bin"

if os.path.exists(model_path):
    print(f"File already exists at {model_path}")
else:
    response = requests.get(model_url, stream=True)
    total_size = int(response.headers.get('content-length', 0))

    # Download the file with progress bar
    with open(model_path, 'wb') as f, tqdm(
        desc=model_path,
        total=total_size,
        unit='iB',
        unit_scale=True,
        unit_divisor=1024,
    ) as bar:
        for data in response.iter_content(chunk_size=1024):
            size = f.write(data)
            bar.update(size)

    if model_path.endswith('.gz'):
        import gzip
        import shutil
        with gzip.open(model_path, 'rb') as f_in:
            with open(extracted_model_path, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

app = FastAPI()

class ResultModel(BaseModel):
    outlier: List[Tuple[int, int]]
    null_values: List[Tuple[int, int]]
    odd_one_outs: List[Tuple[int, int]]

class SheetModel(BaseModel):
    sheet_name: str
    result: ResultModel

class ResponseModel(BaseModel):
    data: List[SheetModel]

def tuple_encoder(obj):
    if isinstance(obj, list) and len(obj) == 2 and all(isinstance(i, int) for i in obj):
        return tuple(obj)
    elif isinstance(obj, list):
        return [tuple_encoder(i) for i in obj]
    elif isinstance(obj, dict):
        return {k: tuple_encoder(v) for k, v in obj.items()}
    return obj


ft = fasttext.load_model('cc.en.300.bin')
# model = KeyedVectors.load_word2vec_format('GoogleNews-vectors-negative300.bin.gz', binary=True)


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def extract_features(values):
    length = len(values)
    num_digits = len(re.findall(r'\d', values))
    num_letters = len(re.findall(r'[A-Za-z]', values))
    num_special = len(re.findall(r'[^A-Za-z0-9]', values))
    return [length, num_digits, num_letters, num_special]

def detect_outlier(df):
    
    try:
        outliers = []
        for i,column in enumerate(df.columns):
            if is_numeric_dtype(df[column]):
                vals = np.sort(df[column].fillna(method='ffill').fillna(method='bfill'))
                q1 = np.percentile(vals, 25, method = 'midpoint') 
                q3 = np.percentile(vals, 75, method = 'midpoint') 

                iqr = q3-q1

                low_lim = q1 - (1.5 * iqr)
                up_lim = q3 + (1.5 * iqr)

                for j,val in enumerate(df[column]):
                    if val<low_lim or val>up_lim:
                        outliers.append((tuple((int(j+1),int(i)))))
                print(up_lim)
        print(outliers)
        return outliers
    except Exception as e:
        print(e)
        raise Exception("Error in Detect Outliers")
    

def detect_null(df):
    try:
        nulls = []
        for i,column in enumerate(df.columns):
            for j in np.where(df[column].isnull())[0]:
                nulls.append((tuple((int(j+1),int(i)))))
        # print(type(nulls),type(nulls[0]),type(nulls[0][0]))
        # print(nulls)
        return nulls
    except Exception as e:
        print(e)
        raise Exception("Error in Detect Null")

def detect_odd_one_out(df,model):
    
    try:
        onos = []
        for i,column in enumerate(df.columns):
            if i == 0:continue
            if not is_numeric_dtype(df[column]):
                words = list(df[column])
                # words = [word.replace(" ","") for word in words]
                word_vectors = [model[word] for word in words if word in model]
                # print(len(word_vectors) , len(words))
                if len(word_vectors) < len(words)/2:
                    print("Using Isolation Forest")
                    features = np.array([extract_features(word) for word in words])
                    iso_model = IsolationForest(contamination=0.25)  # Adjust contamination as needed
                    iso_model.fit(features)
                    predictions = iso_model.predict(features)
                     
                    for j in range(len(predictions)):
                        if predictions[i] == -1:
                            onos.append((tuple((int(j+1),int(i)))))   
                    odd_one_out_words = [words[i] for i in range(len(predictions)) if predictions[i] == -1]
                    
                else:
                    print("Using fasttext")
                    similarities = []
                    word_vectors = []
                    for word in words:
                        if type(word)!=str: continue
                        if word not in ft:
                            word_vectors.append(ft.get_word_vector(word))
                        else:
                            word_vectors.append(ft[word])

                    for index, word_vector in enumerate(word_vectors):
                        # Compute similarity of the word with all other words
                        avg_similarity = sum([1 - cosine(word_vector, other_vector)
                                            for j, other_vector in enumerate(word_vectors) if index != j]) / (len(word_vectors) - 1)
                        similarities.append(avg_similarity)
                    print(similarities)
                    zscores = zscore(similarities)
                    q1 = np.percentile(zscores, 2.5, method = 'midpoint') 
                    print(q1)
                    odd_one_out_index = np.where(zscores < q1)[0]
                    odd_one_out_words = [words[index] for index in odd_one_out_index]
                    for j,val in enumerate(df[column]):
                        if val in odd_one_out_words:
                            onos.append(tuple((int(j+1),int(i)))) 

                print(f"The odd one out is: {odd_one_out_words}")
                return onos
    except Exception as e:
        print(e)
        raise Exception("Error in Detect Odd ones out.")

app = FastAPI()

@app.post("/upload_excel",response_model=ResponseModel)
async def upload_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        data = BytesIO(contents)
        file_name = file.filename
        output = []

        if file_name[-4:] == '.csv':
            df_dict = {"Sheet1":pd.read_csv(data)}
        elif file_name[-5:] == '.xlsx': 
            sheet_names = pd.ExcelFile(data).sheet_names
            print(sheet_names,type(sheet_names))
            df_dict = pd.read_excel(data,sheet_name = sheet_names)
        else:
            return {"message": "File type incorrect."}

        for key in df_dict.keys():
            df = df_dict[key]
            nulls = detect_null(df)
            outliers = detect_outlier(df)
            odd_one_outs = detect_odd_one_out(df,ft)

            output.append({
                "sheet_name":key,
                "result": {
                    "outlier":outliers,
                    "null_values":nulls,
                    "odd_one_outs":odd_one_outs
                }
            })

        return JSONResponse(content=tuple_encoder({"data":output}))
    except Exception as e:
        print(e)
        return {"message": e}
    finally:
        file.file.close()


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)